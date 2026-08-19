from __future__ import annotations

from pathlib import Path
from urllib.parse import quote_plus

from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook

APP_TITLE = "Consumables Marketplace Search Helper"
WORKBOOK_PATH = Path(__file__).with_name("Consumables_Master.xlsx")

app = FastAPI(title=APP_TITLE)


def load_keywords() -> dict[str, list[str]]:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    data: dict[str, list[str]] = {}

    for ws in wb.worksheets:
        if ws.title == "Review Guide":
            continue
        if ws["D1"].value != "Compliance Monitoring Keywords":
            continue

        keywords: list[str] = []
        for row in range(4, ws.max_row + 1):
            value = ws.cell(row=row, column=4).value
            if value:
                keywords.append(str(value).strip())
        data[ws.title] = keywords

    return data


def walmart_search_url(keyword: str) -> str:
    return f"https://www.walmart.com/search?q={quote_plus(keyword)}"


def html_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def render_page(selected_sheet: str | None, filter_text: str | None) -> str:
    data = load_keywords()
    sheets = list(data.keys())
    active_sheet = selected_sheet if selected_sheet in data else (sheets[0] if sheets else "")
    active_keywords = data.get(active_sheet, [])
    active_filter = (filter_text or "").strip().lower()

    filtered_keywords = [
        kw for kw in active_keywords
        if not active_filter or active_filter in kw.lower()
    ]

    nav_items = []
    for sheet in sheets:
        checked = "active" if sheet == active_sheet else ""
        nav_items.append(
            f'<a class="tab {checked}" href="/?sheet={quote_plus(sheet)}">{html_escape(sheet)}</a>'
        )

    rows = []
    for idx, kw in enumerate(filtered_keywords, start=1):
        url = walmart_search_url(kw)
        rows.append(
            "<tr>"
            f'<td><input type="checkbox" class="kw-box" name="kw" form="combined-form" value="{html_escape(kw)}" data-url="{html_escape(url)}" aria-label="Select {html_escape(kw)}" onchange="updateCombinedPreview()"></td>'
            f'<td class="index">{idx}</td>'
            f'<td class="keyword">{html_escape(kw)}</td>'
            f'<td><a href="{html_escape(url)}" target="_blank" rel="noopener noreferrer">Open search</a></td>'
            "</tr>"
        )

    if not rows:
        rows.append('<tr><td colspan="4" class="empty">No keywords match that filter.</td></tr>')

    selected_count = len(filtered_keywords)
    sheet_options = "".join(
        f'<option value="{html_escape(sheet)}" {"selected" if sheet == active_sheet else ""}>{html_escape(sheet)}</option>'
        for sheet in sheets
    )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(APP_TITLE)}</title>
  <style>
    :root {{
      --bg: #f7f8fa;
      --panel: #ffffff;
      --border: #d8dee9;
      --text: #102133;
      --muted: #5e6b7a;
      --accent: #1f4e78;
      --accent-2: #e2f0d9;
      --link: #0b63ce;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: var(--bg); color: var(--text); }}
    header {{ background: linear-gradient(135deg, var(--accent), #2c6aa0); color: white; padding: 20px 24px; }}
    header h1 {{ margin: 0 0 6px 0; font-size: 22px; }}
    header p {{ margin: 0; color: rgba(255,255,255,.9); }}
    main {{ padding: 20px; max-width: 1450px; margin: 0 auto; }}
    .toolbar, .tabs, .card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 12px; box-shadow: 0 1px 2px rgba(0,0,0,.04); }}
    .toolbar {{ padding: 14px; margin-bottom: 16px; display: grid; grid-template-columns: 1fr auto auto auto auto; gap: 12px; align-items: center; }}
    .toolbar label {{ font-size: 13px; color: var(--muted); display: block; margin-bottom: 4px; }}
    .toolbar input, .toolbar select {{ width: 100%; padding: 10px 12px; border: 1px solid var(--border); border-radius: 10px; font-size: 14px; }}
    .toolbar button {{ padding: 10px 14px; border: 0; border-radius: 10px; background: var(--accent); color: white; cursor: pointer; font-weight: 600; }}
    .toolbar button.secondary {{ background: #6b7280; }}
    .tabs {{ padding: 10px; margin-bottom: 16px; display: flex; gap: 8px; flex-wrap: wrap; }}
    .tab {{ text-decoration: none; color: var(--accent); border: 1px solid var(--border); padding: 8px 12px; border-radius: 999px; background: #fff; font-size: 13px; }}
    .tab.active {{ background: var(--accent); color: white; border-color: var(--accent); }}
    .card {{ padding: 16px; }}
    .stats {{ display: flex; gap: 18px; flex-wrap: wrap; margin-bottom: 12px; color: var(--muted); font-size: 14px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 10px 8px; border-bottom: 1px solid var(--border); vertical-align: top; }}
    th {{ text-align: left; font-size: 13px; color: var(--muted); }}
    td.keyword {{ font-weight: 600; }}
    td.index {{ width: 56px; color: var(--muted); }}
    td a {{ color: var(--link); text-decoration: none; }}
    td a:hover {{ text-decoration: underline; }}
    tr:hover td {{ background: #fafcff; }}
    .controls {{ display: flex; gap: 8px; margin: 12px 0 14px 0; flex-wrap: wrap; }}
    .controls button {{ padding: 9px 12px; border-radius: 10px; border: 1px solid var(--border); background: #fff; cursor: pointer; }}
    .controls button.primary {{ background: var(--accent); color: white; border-color: var(--accent); }}
    .empty {{ color: var(--muted); padding: 20px 8px; text-align: center; }}
    .note {{ margin-top: 12px; color: var(--muted); font-size: 13px; line-height: 1.5; }}
    .pill {{ display: inline-block; background: var(--accent-2); border: 1px solid #c8dfb8; color: #27462b; padding: 4px 8px; border-radius: 999px; font-size: 12px; margin-left: 8px; }}
    @media (max-width: 900px) {{
      .toolbar {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
<header>
  <h1>{html_escape(APP_TITLE)}</h1>
  <p>Generate Walmart Marketplace searches from your workbook keywords without manually typing every query. Bless automation.</p>
</header>
<main>
  <form class="toolbar" method="get">
    <div>
      <label for="sheet">Subcategory</label>
      <select id="sheet" name="sheet">{sheet_options}</select>
    </div>
    <div>
      <label for="q">Filter keywords</label>
      <input id="q" name="q" value="{html_escape(filter_text or '')}" placeholder="Type to narrow the list">
    </div>
    <div style="align-self:end;">
      <button type="submit">Refresh</button>
    </div>
    <div style="align-self:end;">
      <button type="button" class="secondary" onclick="clearFilter()">Clear filter</button>
    </div>
    <div style="align-self:end;">
      <button type="button" onclick="copySelected()">Copy selected URLs</button>
    </div>
  </form>

  <div class="tabs">{''.join(nav_items)}</div>

  <div class="card">
    <div class="stats">
      <div><strong>{html_escape(active_sheet)}</strong> <span class="pill">{selected_count} keywords</span></div>
      <div>Use the checkboxes to open or copy batches of search links.</div>
    </div>

    <form id="combined-form" action="/combined-search" method="get" target="_blank"></form>
    <div class="controls">
      <button id="combined-button" class="primary" type="submit" form="combined-form">Search selected together</button>
      <button type="button" onclick="selectAllVisible(true)">Select all visible</button>
      <button type="button" onclick="selectAllVisible(false)">Clear selected</button>
      <button type="button" onclick="openSelected()">Open selected separately</button>
      <button type="button" onclick="copySelected()">Copy selected</button>
    </div>
    <div class="note" id="combo-preview" style="margin-top:-4px;">Select two or more keywords and the combined search will update automatically.</div>

    <table>
      <thead>
        <tr>
          <th style="width:40px;"><input type="checkbox" id="master" onchange="toggleMaster(this)"></th>
          <th>#</th>
          <th>Keyword</th>
          <th>Walmart search</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows)}
      </tbody>
    </table>

    <div class="note">
      Tip: use the filter box to narrow down one theme at a time, then open or copy the batch. This is a local helper only — no ad group, no pipeline drama.
    </div>
  </div>
</main>
<script>
  function visibleBoxes() {{
    return Array.from(document.querySelectorAll('.kw-box'));
  }}
  function toggleMaster(el) {{
    visibleBoxes().forEach(cb => cb.checked = el.checked);
    updateCombinedPreview();
  }}
  function selectAllVisible(on) {{
    visibleBoxes().forEach(cb => cb.checked = on);
    document.getElementById('master').checked = on;
    updateCombinedPreview();
  }}
  function selectedKeywords() {{
    return visibleBoxes().filter(cb => cb.checked).map(cb => cb.value);
  }}
  function selectedUrls() {{
    return visibleBoxes().filter(cb => cb.checked).map(cb => cb.dataset.url || cb.value);
  }}
  function updateCombinedPreview() {{
    const button = document.getElementById('combined-button');
    const preview = document.getElementById('combo-preview');
    const keywords = selectedKeywords();
    button.disabled = keywords.length === 0;
    preview.textContent = keywords.length
      ? 'Combined query: ' + keywords.join(' ')
      : 'Select two or more keywords and the combined search will update automatically.';
  }}
  async function copySelected() {{
    const urls = selectedUrls();
    if (!urls.length) {{ alert('Pick at least one keyword first.'); return; }}
    try {{
      await navigator.clipboard.writeText(urls.join('\n'));
      alert('Copied ' + urls.length + ' URLs to clipboard.');
    }} catch (err) {{
      alert('Clipboard copy failed. You can still open links manually.');
    }}
  }}
  function clearFilter() {{
    const q = document.getElementById('q');
    q.value = '';
    document.getElementById('sheet').form.submit();
  }}
  window.addEventListener('DOMContentLoaded', updateCombinedPreview);
</script>
</body>
</html>"""


@app.get('/combined-search')
def combined_search(kw: list[str] = Query(default=[])):
    if not kw:
        return RedirectResponse(url='/', status_code=302)
    combined = ' '.join(kw)
    return RedirectResponse(url=walmart_search_url(combined), status_code=302)



@app.get('/', response_class=HTMLResponse)
def index(
    sheet: str | None = Query(default=None),
    q: str | None = Query(default=None, alias='q'),
):
    return HTMLResponse(render_page(sheet, q))


if __name__ == '__main__':
    import uvicorn

    uvicorn.run(app, host='127.0.0.1', port=8000, reload=False)
